import os
import json
import re
import time
import streamlit as st
import google.genai as genai
import io

st.set_page_config(
    page_title="네이버 검색광고 키워드 추출기",
    page_icon="🔍",
    layout="wide",
)

FALLBACK_MODELS = ['gemini-2.0-flash', 'gemini-1.5-flash', 'gemini-2.5-flash']

FEW_SHOT_EXAMPLES = """
[예시 1]
브랜드: 마인드브릿지
상품: [올데이] 옥스포드 오버핏 셔츠_4colors / [올데이] 데님 오버핏 셔츠_3colors / [COUTURE] 와이드 원턱 밴딩 팬츠_2colors / 테이퍼드 밴딩 팬츠_2colors / 스웻하프팬츠_2colors / [올데이] 수피마 오버핏 셔츠_3colors
추출 (27개): 마인드브릿지데님셔츠, 마인드브릿지데님오버핏셔츠, 마인드브릿지세미와이드, 마인드브릿지세미와이드밴딩팬츠, 마인드브릿지수피마, 마인드브릿지수피마셔츠, 마인드브릿지수피마오버핏셔츠, 마인드브릿지스웻, 마인드브릿지스웻팬츠, 마인드브릿지스웻하프팬츠, 마인드브릿지슬랙스, 마인드브릿지오버핏데님셔츠, 마인드브릿지오버핏수피마셔츠, 마인드브릿지옴브레체크, 마인드브릿지옴브레체크셔츠, 마인드브릿지와이드슬랙스, 마인드브릿지와이드원턱팬츠, 마인드브릿지와이드팬츠, 마인드브릿지원턱팬츠, 마인드브릿지하프팬츠

[예시 2]
브랜드: 아이오페
상품: 슈퍼바이탈 아이크림 25ml / 슈퍼바이탈 크림 세트 / 슈퍼바이탈 2종 (소프너+에멀젼) 세트 / 비타민C 25% 항산화 토닝앰플 2개 / UV 쉴드 선 프로텍터 2개 / UV 쉴드 톤업 선 2개 / 레티놀 엑스퍼트 0.1% 기획세트
추출 (16개): 아이오페, 아이오페레티놀, 아이오페레티놀엑스퍼트, 아이오페비타민, 아이오페선크림, 아이오페소프너, 아이오페슈퍼바이탈, 아이오페슈퍼바이탈크림, 아이오페아이크림, 아이오페앰플, 아이오페에멀전, 아이오페에센스, 아이오페오일, 아이오페크림, 아이오페토닝앰플, 아이오페톤업선크림
"""

SYSTEM_PROMPT = """네이버 검색광고 키워드 추출 전문가. 브랜드+상품목록 → 검색광고용 키워드 추출.

## 규칙
1. 형식: 브랜드명+핵심단어 결합. 띄어쓰기/특수문자 절대 불포함.
2. 상품명에서 핵심 단어 추출 (대괄호, 색상, 용량, 수량, 할인율 제거. 영문→한글 변환/제거)
3. 브랜드+단어1, 브랜드+단어1+단어2 등 다양한 조합 생성. 연관 단어 포함 (롱슬리브→긴팔 등)
4. 브랜드 단독 키워드 포함 (예: 아이오페)
5. **브랜드당 20~30개** 키워드만 추출 (핵심적이고 검색 가능성 높은 키워드 우선)
6. 중복 제거, 가나다순 정렬
7. 필수키워드가 제공되면: 결과에서 제외하되, 패턴/방향성을 참고하여 확장 추출

## 절대 금지 (매우 중요)
- **브랜드명 중복 금지**: 키워드 안에 브랜드명이 두 번 이상 들어가면 안 됩니다.
  - 나쁜 예: 디올디올쇼, 아이오페아이오페크림 → 브랜드명이 두 번 반복됨
  - 좋은 예: 디올립글로우, 아이오페크림 → 브랜드명 1번 + 상품키워드
- **의미 없는 단어 조합 금지**: 브랜드명 뒤에 붙이는 단어는 반드시 소비자가 실제 검색할 만한 **상품 카테고리/종류**(셔츠, 크림, 자켓, 향수, 립스틱 등), **소재**(울, 데님, 수피마 등), **라인명**(슈퍼바이탈, 레티놀 등), 또는 이들의 조합이어야 합니다.
  - 나쁜 예: 마리끌레르버튼, 마리끌레르베이직, 마리끌레르오픈 → 단독으로 의미 없는 수식어
  - 좋은 예: 마리끌레르롱슬리브, 마리끌레르울자켓, 마리끌레르체크셔츠 → 검색 가능한 상품 카테고리 포함

## 예시
""" + FEW_SHOT_EXAMPLES + """
## 응답 형식 (JSON만, 다른 텍스트 없이)
```json
{"브랜드명1": ["키워드1", "키워드2", ...]}
```
"""


def get_api_key():
    if "GEMINI_API_KEY" in st.secrets:
        return st.secrets["GEMINI_API_KEY"]
    return os.getenv("GEMINI_API_KEY", "")


def clean_keywords(keywords, brand_name=None, existing_keywords=None):
    existing = set(existing_keywords or [])
    brand_lower = brand_name.lower().replace(' ', '') if brand_name else None
    cleaned = []
    for kw in keywords:
        kw = re.sub(r'[^가-힣a-zA-Z0-9]', '', kw)
        kw = kw.strip()
        if not kw or len(kw) <= 1 or kw in existing:
            continue
        if brand_lower:
            kw_lower = kw.lower()
            if kw_lower.startswith(brand_lower):
                remainder = kw_lower[len(brand_lower):]
                if remainder.startswith(brand_lower):
                    continue
        cleaned.append(kw)
    return sorted(set(cleaned))


def is_retryable_error(error_str):
    retryable = ['429', 'RESOURCE_EXHAUSTED', '503', 'UNAVAILABLE', '500', 'INTERNAL']
    return any(code in error_str for code in retryable)


def parse_keyword_text(text):
    return [k.strip() for k in re.split(r'[;,\n]+', text) if k.strip()]


def extract_keywords_api(brands, api_key):
    client = genai.Client(api_key=api_key)
    all_existing = {}

    user_prompt = "아래 브랜드와 상품 목록에 대해 네이버 검색광고용 키워드를 추출해주세요.\n\n"
    for brand_info in brands:
        brand_name = brand_info['brand']
        products = brand_info['products']
        existing_kw = brand_info.get('existing', [])

        existing_clean = [re.sub(r'[^가-힣a-zA-Z0-9]', '', k).strip() for k in existing_kw if k.strip()]
        existing_clean = [k for k in existing_clean if k]
        all_existing[brand_name] = set(existing_clean)

        user_prompt += f"브랜드: {brand_name}\n상품목록:\n"
        for p in products:
            user_prompt += f"- {p}\n"

        if existing_clean:
            user_prompt += f"\n필수키워드 (이미 등록됨, 결과에서 제외하되 방향성 참고):\n"
            for ek in existing_clean:
                user_prompt += f"- {ek}\n"
        user_prompt += "\n"

    last_error = None
    for model_name in FALLBACK_MODELS:
        for attempt in range(2):
            try:
                response = client.models.generate_content(
                    model=model_name,
                    contents=user_prompt,
                    config={
                        'system_instruction': SYSTEM_PROMPT,
                        'temperature': 0.3,
                    }
                )
                response_text = response.text.strip()
                json_match = re.search(r'\{[\s\S]*\}', response_text)
                if not json_match:
                    last_error = "AI 응답 파싱 실패"
                    break

                result = json.loads(json_match.group())
                for bn in result:
                    existing = all_existing.get(bn, set())
                    result[bn] = clean_keywords(result[bn], bn, existing)
                return result, model_name

            except Exception as e:
                last_error = str(e)
                if is_retryable_error(last_error) and attempt < 1:
                    time.sleep(5)
                    continue
                else:
                    break

    raise Exception(f"모든 AI 모델이 일시적으로 사용 불가합니다. 잠시 후 다시 시도해주세요.\n({last_error[:200] if last_error else ''})")


# ── Custom CSS ──

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Noto Sans KR', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    .block-container {
        max-width: 960px;
        padding-top: 0 !important;
    }

    /* Header */
    .app-header {
        background: linear-gradient(135deg, #03c75a 0%, #00a550 100%);
        color: white;
        padding: 2.5rem 2rem 2rem;
        border-radius: 0 0 20px 20px;
        margin: -1rem -1rem 2rem -1rem;
        text-align: center;
        box-shadow: 0 8px 32px rgba(3, 199, 90, 0.2);
    }
    .app-header h1 {
        font-size: 2rem;
        font-weight: 800;
        margin: 0 0 6px 0;
        letter-spacing: -0.5px;
    }
    .app-header p {
        font-size: 0.85rem;
        opacity: 0.85;
        margin: 0;
        font-weight: 400;
    }

    /* Brand section */
    .brand-section-title {
        display: inline-flex;
        align-items: center;
        gap: 10px;
        margin: 4px 0 10px 0;
    }
    .brand-num {
        background: #03c75a;
        color: white;
        width: 26px;
        height: 26px;
        border-radius: 50%;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-weight: 700;
        font-size: 12px;
        flex-shrink: 0;
    }
    .brand-label {
        font-weight: 700;
        font-size: 15px;
        color: #1a1a1a;
    }

    /* Usage tip */
    .usage-tip {
        background: #f8fafb;
        border-left: 3px solid #03c75a;
        padding: 12px 16px;
        border-radius: 0 8px 8px 0;
        font-size: 13px;
        color: #495057;
        line-height: 1.7;
        margin-bottom: 1.5rem;
    }
    .usage-tip strong {
        color: #1a1a1a;
    }
    .field-label {
        font-size: 12px;
        font-weight: 600;
        color: #868e96;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 6px;
    }
    .field-label-kw {
        font-size: 12px;
        font-weight: 600;
        color: #e67e22;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 6px;
    }
    .count-badge {
        display: inline-block;
        background: #f1f3f5;
        color: #495057;
        padding: 2px 10px;
        border-radius: 20px;
        font-size: 11px;
        font-weight: 600;
        margin-left: 4px;
    }

    /* Stat cards */
    .stat-card {
        background: #ffffff;
        border: 1px solid #eaecef;
        border-radius: 14px;
        padding: 1.25rem 1rem;
        text-align: center;
        box-shadow: 0 1px 4px rgba(0,0,0,0.04);
    }
    .stat-val {
        font-size: 2rem;
        font-weight: 800;
        color: #03c75a;
        line-height: 1.2;
    }
    .stat-lbl {
        font-size: 0.75rem;
        color: #868e96;
        font-weight: 500;
        margin-top: 2px;
    }

    /* Result brand */
    .result-brand {
        background: #ffffff;
        border: 1px solid #eaecef;
        border-radius: 14px;
        overflow: hidden;
        margin-bottom: 1rem;
        box-shadow: 0 1px 4px rgba(0,0,0,0.04);
    }
    .result-brand-head {
        background: linear-gradient(135deg, #f0fdf5 0%, #e8f8ef 100%);
        padding: 14px 20px;
        display: flex;
        justify-content: space-between;
        align-items: center;
        border-bottom: 1px solid #eaecef;
    }
    .result-brand-head strong {
        font-size: 15px;
        color: #1a1a1a;
    }
    .kw-count-pill {
        background: #03c75a;
        color: white;
        padding: 4px 14px;
        border-radius: 20px;
        font-size: 12px;
        font-weight: 700;
    }
    .kw-grid {
        padding: 16px 20px;
        display: flex;
        flex-wrap: wrap;
        gap: 6px;
    }
    .kw-chip {
        display: inline-block;
        background: #f8f9fa;
        border: 1px solid #e9ecef;
        color: #343a40;
        padding: 7px 14px;
        border-radius: 8px;
        font-size: 13px;
        font-weight: 500;
        transition: all 0.15s;
        cursor: default;
    }
    .kw-chip:hover {
        background: #e8f8ef;
        border-color: #03c75a;
        color: #02a94d;
    }

    /* Hide streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    /* Fix textarea */
    textarea {
        font-family: 'Noto Sans KR', sans-serif !important;
        font-size: 13px !important;
    }

    /* Streamlit buttons override */
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #03c75a 0%, #00a550 100%) !important;
        border: none !important;
        font-weight: 700 !important;
        font-size: 15px !important;
        padding: 0.7rem 2rem !important;
        border-radius: 12px !important;
        box-shadow: 0 4px 16px rgba(3, 199, 90, 0.3) !important;
        transition: all 0.2s !important;
    }
    .stButton > button[kind="primary"]:hover {
        box-shadow: 0 6px 24px rgba(3, 199, 90, 0.4) !important;
        transform: translateY(-1px);
    }
    .stButton > button[kind="secondary"] {
        border-radius: 10px !important;
        font-weight: 600 !important;
        font-size: 13px !important;
    }
    .stDownloadButton > button {
        border-radius: 10px !important;
        font-weight: 600 !important;
    }

    /* Divider */
    hr {
        border: none;
        border-top: 1px solid #f1f3f5;
        margin: 0.5rem 0;
    }
</style>
""", unsafe_allow_html=True)


# ── Header ──

st.markdown("""
<div class="app-header">
    <h1>네이버 검색광고 키워드 추출기</h1>
    <p>브랜드와 상품명을 입력하면 AI가 검색광고용 키워드를 자동 추출합니다</p>
</div>
""", unsafe_allow_html=True)


# ── API Key ──

api_key = get_api_key()
if not api_key:
    st.warning("Gemini API Key가 설정되지 않았습니다. 아래에 입력하거나, Streamlit Secrets에 `GEMINI_API_KEY`를 등록하세요.")
    api_key = st.text_input("Gemini API Key", type="password")
    if not api_key:
        st.stop()

st.markdown("""
<div class="usage-tip">
    <strong>사용법</strong><br>
    브랜드명을 입력하고, 상품명을 한 줄에 하나씩 입력하세요. 상품 개수 제한 없이 자유롭게 등록 가능합니다.<br>
    이미 운영 중인 필수키워드가 있다면 함께 입력하면 중복 없이 확장된 키워드를 추출합니다.
</div>
""", unsafe_allow_html=True)

if 'brand_count' not in st.session_state:
    st.session_state.brand_count = 1


# ── Brand Inputs ──

brands_data = []
for i in range(st.session_state.brand_count):
    st.markdown(f"""
    <div class="brand-section-title">
        <span class="brand-num">{i+1}</span>
        <span class="brand-label">브랜드</span>
    </div>
    """, unsafe_allow_html=True)

    brand_name = st.text_input(
        f"브랜드명 {i+1}",
        key=f"brand_{i}",
        placeholder="브랜드명을 입력하세요 (예: 마리끌레르)",
        label_visibility="collapsed",
    )

    col_left, col_right = st.columns(2)
    with col_left:
        st.markdown('<div class="field-label">상품 목록</div>', unsafe_allow_html=True)
        products_text = st.text_area(
            f"상품목록 {i+1}",
            key=f"products_{i}",
            height=160,
            placeholder="한 줄에 하나씩 입력\n\n울 버튼 핸드메이드 자켓\n아치 로고 래글런 롱슬리브\n베이직 로고 오버핏 미니 체크 셔츠",
            label_visibility="collapsed",
        )
    with col_right:
        st.markdown('<div class="field-label-kw">필수키워드 (선택)</div>', unsafe_allow_html=True)
        existing_text = st.text_area(
            f"필수키워드 {i+1}",
            key=f"existing_{i}",
            height=160,
            placeholder="이미 등록된 키워드 붙여넣기\n세미콜론·쉼표·줄바꿈 모두 지원\n\n불가리;불가리퍼퓸;불가리퍼퓸오드퍼퓸",
            label_visibility="collapsed",
        )

    products = [p.strip() for p in products_text.split('\n') if p.strip()]
    existing = parse_keyword_text(existing_text) if existing_text else []

    prod_count = len(products)
    ek_count = len(existing)
    st.markdown(
        f'<span class="count-badge">상품 {prod_count}개</span>'
        f'<span class="count-badge">필수키워드 {ek_count}개</span>',
        unsafe_allow_html=True,
    )

    brands_data.append({
        'brand': brand_name.strip() if brand_name else '',
        'products': products,
        'existing': existing,
    })

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)


# ── Action Buttons (브랜드 추가/삭제 → 키워드 추출 순서) ──

col_add, col_del, col_spacer = st.columns([1, 1, 2])
with col_add:
    if st.button("＋ 브랜드 추가", use_container_width=True):
        st.session_state.brand_count += 1
        st.rerun()
with col_del:
    if st.session_state.brand_count > 1:
        if st.button("－ 브랜드 삭제", use_container_width=True):
            st.session_state.brand_count -= 1
            st.rerun()

st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

if st.button("🔍  키워드 추출", type="primary", use_container_width=True):
    valid_brands = [b for b in brands_data if b['brand'] and b['products']]

    if not valid_brands:
        st.error("브랜드명과 상품을 최소 1개 이상 입력해주세요.")
    else:
        with st.spinner("AI가 최적의 키워드를 분석하고 있습니다..."):
            try:
                result, model_used = extract_keywords_api(valid_brands, api_key)
                st.session_state.last_result = result
                st.session_state.model_used = model_used
            except Exception as e:
                st.error(str(e))
                st.stop()


# ── Results ──

if 'last_result' in st.session_state:
    result = st.session_state.last_result

    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

    brand_names = list(result.keys())
    total_kw = sum(len(result[b]) for b in brand_names)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f'<div class="stat-card"><div class="stat-val">{len(brand_names)}</div><div class="stat-lbl">브랜드 수</div></div>', unsafe_allow_html=True)
    with col2:
        st.markdown(f'<div class="stat-card"><div class="stat-val">{total_kw}</div><div class="stat-lbl">총 키워드 수</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    all_keywords_text = ""
    csv_data = "\ufeff브랜드,키워드\n"

    for brand_name in brand_names:
        keywords = result[brand_name]
        all_keywords_text += "\n".join(keywords) + "\n"
        for kw in keywords:
            csv_data += f"{brand_name},{kw}\n"

        chips_html = "".join(f'<span class="kw-chip">{kw}</span>' for kw in keywords)
        st.markdown(f"""
        <div class="result-brand">
            <div class="result-brand-head">
                <strong>{brand_name}</strong>
                <span class="kw-count-pill">{len(keywords)}개</span>
            </div>
            <div class="kw-grid">{chips_html}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    dl1, dl2, dl3 = st.columns(3)
    with dl1:
        st.download_button(
            "📋  TXT 다운로드",
            data=all_keywords_text,
            file_name="naver_keywords.txt",
            mime="text/plain",
            use_container_width=True,
        )
    with dl2:
        st.download_button(
            "📊  CSV 다운로드",
            data=csv_data,
            file_name="naver_keywords.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with dl3:
        try:
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)
            for bn in brand_names:
                ws = wb.create_sheet(title=bn[:31])
                ws.append(["키워드"])
                ws.column_dimensions['A'].width = 40
                for kw in result[bn]:
                    ws.append([kw])
            ws_all = wb.create_sheet(title="전체")
            ws_all.append(["브랜드", "키워드"])
            ws_all.column_dimensions['A'].width = 20
            ws_all.column_dimensions['B'].width = 40
            for bn in brand_names:
                for kw in result[bn]:
                    ws_all.append([bn, kw])
            buf = io.BytesIO()
            wb.save(buf)
            st.download_button(
                "📥  Excel 다운로드",
                data=buf.getvalue(),
                file_name="naver_keywords.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except ImportError:
            st.caption("Excel 다운로드에는 openpyxl 패키지가 필요합니다.")
